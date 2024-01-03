import os
import time
import pyperclip
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
import pyautogui
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Constants
CHATGPT_URL = "https://chat.openai.com/"
WORDPRESS_URL = "https://zingmind.com/wp-admin/post-new.php"

EXCEL_PATH = "C:/Users/hp/Documents/blog.xlsx"
IMAGE_SAVE_PATH = "C:\\Users\\hp\\Downloads\\image.jpg"

# Load environment variables
load_dotenv("info.env")
EMAIL = os.getenv('CHATGPT_EMAIL')
PASSWORD = os.getenv('CHATGPT_PASSWORD')
WORD_USERNAME = os.getenv('WORD_USERNAME')
WORD_PASSWORD = os.getenv('WORD_PASSWORD')
user_input = input("Enter 0 to execute ChatGPT or 1 for another platform: ")

options = uc.ChromeOptions()
driver = uc.Chrome(options=options)
wait = WebDriverWait(driver, 100)
options.add_argument("--headless")


def login_chatgpt():
    driver.get(CHATGPT_URL)

    # driver.get("https://chat.openai.com/")
    sign_in_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="relative flex h-12 items-center justify-center rounded-md text-center text-base font-medium bg-[#3C46FF] text-[#fff] hover:bg-[#0000FF]"]')))
    sign_in_button.click()
    time.sleep(2)

    email_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='username']")))
    email_field.send_keys(EMAIL)
    time.sleep(2)

    continue_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@type='submit']")))
    continue_button.click()
    time.sleep(2)

    password_field = wait.until(EC.presence_of_element_located((By.ID, "password")))
    password_field.send_keys(PASSWORD)
    time.sleep(2)

    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, '_button-login-password')]")))
    element.click()
    time.sleep(2)



def login_wordpress(driver):
    driver.get(WORDPRESS_URL)
    # Implement your login logic for WordPress
    username_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='user_login']")))
    username_field.send_keys(WORD_USERNAME)
    time.sleep(2)
    password_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='user_pass']")))
    password_field.send_keys(WORD_PASSWORD)
    time.sleep(2)

    login_button = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='submit']")))
    login_button.click()
    time.sleep(2)

def fetch_content_from_chatgpt(title):
    # #Implement logic to fetch content from ChatGPT based on title
    driver.get(CHATGPT_URL)
   
    search_field1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//textarea[@id='prompt-textarea']")))
    search_field1.clear()
    search_query = f"write a technical blog on this topic with 200 words and include an introduction and conclusion - {title}"
    search_field1.send_keys(search_query)
    time.sleep(5)
    search_field1.send_keys(Keys.ENTER)
    time.sleep(10)
    # Send search query to ChatGPT
  

    # # Wait for the response
    response_element = WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@class, 'markdown')]")))
    content = response_element.get_attribute("outerHTML")

    # Save the output to a text file
    output_file_path = "C:\\Users\\hp\\Downloads\\blog1.txt"
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        output_file.write(content)

    time.sleep(5)
    driver.get("https://www.chatgpt.com/")

    search_field2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//textarea[@id='prompt-textarea']")))
    search_field2.clear()
    # Send search query to ChatGPT
    search_field2.send_keys(f"write a one word keyphara related to  -{title}.")
    search_field2.send_keys(Keys.ENTER)

    time.sleep(5)

    keyphrase_content = WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.markdown.prose.w-full.break-words.dark\:prose-invert.light')))
    content1 = keyphrase_content.text

    # Save the output to a text file
    output_file_path1 = "C:\\Users\\hp\\Downloads\\keys1.txt"
    with open(output_file_path1, 'w', encoding='utf-8') as output_file:
        output_file.write(content1)

    driver.get("https://www.chatgpt.com/")

    # Send search query to ChatGPT
    search_field3 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//textarea[@id='prompt-textarea']")))
    search_field3.clear()
    search_field3.send_keys(f"write the top 10 tags only headings related to  -{title}.")
    search_field3.send_keys(Keys.ENTER)

    time.sleep(5)

    response_element = WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@class, 'markdown')]")))
    content2 = response_element.text

    # Save the output to a text file
    output_file_path2 = "C:\\Users\\hp\\Downloads\\tags1.txt"
    with open(output_file_path2, 'w', encoding='utf-8') as output_file:
        output_file.write(content2)

   
    time.sleep(5)





def post_content_to_wordpress(title, content, keyphrase, tag_column, category, imagename,meta_dis):
    # Implement logic to post content to WordPress
    driver.get("https://www.google.com") 
    time.sleep(2)

    textinput=driver.find_element(By.NAME,"q")
    textinput.send_keys(f"free watermark images with 200*200 pixel for {title}.")

    textinput.send_keys(Keys.RETURN)
    time.sleep(5)

    images_link = driver.find_element(By.LINK_TEXT, "Images")

    images_link.click()
    time.sleep(5)

    image_element = driver.find_element(By.CSS_SELECTOR, "img.rg_i.Q4LuWd")
    image_element.click()

    time.sleep(5)
        # Right-click on the image using the Actions class
    actions = ActionChains(driver)
    actions.context_click(image_element).perform()
    time.sleep(2)
        # Instead of actions, use pyautogui to navigate context menu and save
    pyautogui.press('down', presses=6)  # Press the down arrow key 3 times
    pyautogui.press('enter')  # Press the 'Enter' key

    time.sleep(5)
    # Switch to the new tab
    driver.switch_to.window(driver.window_handles[1])

    # Get the URL of the current page (which is the image URL)
    image_url = driver.current_url
    pyperclip.copy(image_url)

    time.sleep(5)
    driver.get(WORDPRESS_URL)
    time.sleep(5)
            
    feature_image=driver.find_element(By.XPATH,"//a[@id='set-post-thumbnail']")
    feature_image.click()
    time.sleep(5)

    upload_button = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//button[@id='menu-item-upload']")))
    upload_button.click()
        # upload_button.click()

    time.sleep(5)

    select_file=WebDriverWait(driver,30).until(EC.visibility_of_element_located((By.XPATH,"//button[@class='browser button button-hero']")))
    select_file.click()
    time.sleep(5)
    extension = ".jpg" 
    file_path = f"C:\\Users\\hp\\Downloads\\images\\{imagename}{extension}"
    pyautogui.write(file_path)
    pyautogui.press('enter')
    time.sleep(5)

    alt_text=driver.find_element(By.ID,"attachment-details-alt-text")
    alt_text.send_keys(title)
    time.sleep(5)

    caption=driver.find_element(By.ID,"attachment-details-caption")
    caption.send_keys(title)
    time.sleep(5)


    discription=driver.find_element(By.ID,"attachment-details-description")
    discription.send_keys(title)
    time.sleep(10)

    # image_element_media = driver.find_element(By.ID,"menu-item-upload")
    # driver.execute_script("arguments[0].click();", image_element_media)
    # time.sleep(5)

    set_feature_image_button = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "(//button[@class='button media-button button-primary button-large media-button-select'])[1]")))
    set_feature_image_button.click()
    time.sleep(5)


    time.sleep(5)
    insert_media_button = driver.find_element(By.ID, "insert-media-button")
    insert_media_button.click()
    time.sleep(5)
    


    insert_url=driver.find_element(By.ID,"menu-item-embed")

    insert_url.click()
    time.sleep(5)

    insert_url_link = driver.find_element(By.ID, "embed-url-field")
    insert_url_link.click()

    # Simulate Ctrl + V for pasting content
    insert_url_link.send_keys(Keys.CONTROL, 'v')

    time.sleep(5)

    alternative_text=driver.find_element(By.ID,"embed-image-settings-alt-text")
    alternative_text.send_keys(title)
    time.sleep(2)
    caption=driver.find_element(By.ID,"embed-image-settings-caption")

    caption.send_keys(title)


    time.sleep(10)

    button_element = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "(//button[@class='button media-button button-primary button-large media-button-select'])[2]")))

# Perform the click action
    button_element.click()
    time.sleep(5)
    #feature image select
    

   
    time.sleep(2)
    
    
    add_title = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='post_title']")))
    add_title.send_keys(title)
    time.sleep(5)

 
    input_file_paths="C:\\Users\\hp\\Downloads\\blog1.txt"
    with open(input_file_paths, 'r', encoding='utf-8') as input_file:
        content = input_file.read()

    # Wait for the content block to be clickable
    add_content = wait.until(EC.element_to_be_clickable((By.ID, "content")))
    add_content.send_keys(content)

    driver.execute_script("window.scrollBy(0, 500);")
    time.sleep(2)
    
    social_media=driver.find_element(By.ID,"wpseo-meta-tab-social")
    social_media.click()
    time.sleep(2)
    scroll_pixels = 200
    driver.execute_script(f'window.scrollBy(0, {scroll_pixels});')

    facebook_image = driver.find_element(By.ID, "facebook-select-button-metabox")
    facebook_image.click()   
    time.sleep(5)    







    # try:
    #     # Loop through a range of indexes
    #     for i in range(1, 6):  # Change the range as needed
    #         # Find the div element by CSS selector with incremented index
    #         div_element = WebDriverWait(driver, 10).until(
    #             EC.presence_of_element_located((By.CSS_SELECTOR, f".thumbnail:nth-of-type({i})"))
    #         )

    #         # Find the image element within the div
    #         image_element = div_element.find_element(By.TAG_NAME, "img")

    #         # Click on the image using JavaScript
    #         driver.execute_script("arguments[0].click();", image_element)

    # except Exception as e:
    #     print("Error:", e)

    # element = driver.find_element(By.XPATH,"(//button[@class='media-menu-item active'])[1]")
    # driver.execute_script("arguments[0].scrollIntoView(true);", element)
    # time.sleep(10)
    # try:
    #     media_button = WebDriverWait(driver, 30).until(
    #         EC.element_to_be_clickable((By.XPATH, "(//button[@class='media-menu-item'])[1]"))
    #     )
    #     # Hover on the element
    #     media_button=webdriver.ActionChains(driver).move_to_element(media_button).perform()
    #     # Proceed with further actions
    #     # ...
    # except Exception as e:
    #     print(f"Exception occurred: {e}")

    # element = driver.find_element(By.XPATH,"(//button[@class='media-menu-item active'])[1]")
    # driver.execute_script("arguments[0].scrollIntoView(true);", element)
    # time.sleep(10)
    # div_element = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.XPATH, "(//div[@class='centered'])[1]"))
    #     )
    # driver.execute_script("arguments[0].click();", div_element)

        
    upload_buttons = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.XPATH, "//button[@id='menu-item-upload']"))
    )

    # Use the specific index for clicking
    upload_buttons[2].click()  # Assuming you want the second element

        # upload_button.click()

    time.sleep(5)

    select_file=WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.XPATH,"//button[@class='browser button button-hero']")))
    select_file[2].click()
    time.sleep(5)
    extension = ".jpg" 
    file_path = f"C:\\Users\\hp\\Downloads\\images\\{imagename}{extension}"
    pyautogui.write(file_path)
    pyautogui.press('enter')


    time.sleep(10)
    select_button = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "(//button[@class='button media-button button-primary button-large media-button-select'])[3]")))
    select_button.click()


    time.sleep(10)    
    facebook_title=driver.find_element(By.XPATH,"//*[@id='facebook-title-input-metabox']/div/div/div")
    facebook_title.send_keys("datamigration")
    time.sleep(5)
    facebook_discriotion=driver.find_element(By.XPATH,"//*[@id='facebook-description-input-metabox']/div/div")
    facebook_discriotion.send_keys("datamigration")
    time.sleep(10)    
    twitter_image = driver.find_element(By.XPATH, "(//button[@class='yoast-image-select__preview yoast-image-select__preview--no-preview'])[1]")
    twitter_image.click()
    time.sleep(5)
    # div_element = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.XPATH, "(//div[@class='centered'])[2]"))
    #     )
    # driver.execute_script("arguments[0].click();", div_element)



        
    upload_buttons = WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.XPATH, "//button[@id='menu-item-upload']"))
    )

    # Use the specific index for clicking
    upload_buttons[3].click()  # Assuming you want the second element

        # upload_button.click()

    time.sleep(5)

    select_file=WebDriverWait(driver, 30).until(
        EC.presence_of_all_elements_located((By.XPATH,"//button[@class='browser button button-hero']")))
    select_file[3].click()
    time.sleep(5)
    extension = ".jpg" 
    file_path = f"C:\\Users\\hp\\Downloads\\images\\{imagename}{extension}"
    pyautogui.write(file_path)
    pyautogui.press('enter')

    time.sleep(5)
    select_button = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "(//button[@class='button media-button button-primary button-large media-button-select'])[4]")))
    select_button.click()
    
    time.sleep(5)

    twitter_title=driver.find_element(By.XPATH,"//*[@id='twitter-title-input-metabox']/div/div/div")
    twitter_title.send_keys(title)

    time.sleep(5)
    twitter_discribe=driver.find_element(By.XPATH,"//*[@id='twitter-description-input-metabox']/div/div/div")
    twitter_discribe.send_keys(title)
    time.sleep(10)




    driver.execute_script("window.scrollTo(0,800);")
    input_file_path="C:\\Users\\hp\\Downloads\\keys1.txt"
    with open(input_file_path, 'r', encoding='utf-8') as input_file:
        content1 = input_file.read()  
    #facebook image select
        
    
    seo_click=driver.find_element(By.XPATH,"//a[@id='wpseo-meta-tab-content']")
    seo_click.click()
    time.sleep(5)
    keyphrase_input = wait.until(EC.element_to_be_clickable((By.ID, "focus-keyword-input-metabox")))
    keyphrase_input.clear()
    keyphrase_input.send_keys(content1)    

    # time.sleep(5)

    
    meta_discription=driver.find_element(By.XPATH,'//*[@id="yoast-google-preview-description-metabox"]/div/div/div')
    meta_discription.send_keys(meta_dis)
    time.sleep(5)

    driver.execute_script("window.scrollTo(0,500)")
    time.sleep(5)
    category_all_tab = WebDriverWait(driver, 10).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href='#category-all']"))
)
    
    time.sleep(5)
    driver.execute_script("window.scrollBy(0,500)")
    time.sleep(5)

# Click on the category-all tab
    category_all_tab.click()
    time.sleep(5)
    cloud_migration_checkbox2 = driver.find_element(By.XPATH, "//input[@id='in-category-160']")
    cloud_migration_checkbox2.click()
    time.sleep(5)
    cloud_migration_checkbox = driver.find_element(By.XPATH, "//input[@id='in-category-159']")

    cloud_migration_checkbox.click()
    time.sleep(2)



    cloud_migration_checkbox1 = driver.find_element(By.XPATH, "//input[@id='in-category-158']")
    cloud_migration_checkbox1.click()
    time.sleep(5)



    driver.execute_script("window.scrollBy(0, 200);")     
    tags_file_path = "C:\\Users\\hp\\Downloads\\tags1.txt"
    with open(tags_file_path, 'r', encoding='utf-8') as tags_file:
        tags_content = tags_file.read().splitlines()

    for tag in tags_content[1:]:  # Skip the first line assuming it's a header or irrelevant content
        # Fill in the tags field in WordPress
        tags_field = wait.until(EC.element_to_be_clickable((By.ID, "new-tag-post_tag")))
        tags_field.clear()
        tags_field.send_keys(tag)
        time.sleep(1)  # Adjust the sleep duration based on your specific requirements
        
        # Click the "Add" button
        add_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='button tagadd']")))
        add_button.click()

        time.sleep(5)  # Adjust the sleep duration based on your specific requirements
    
        



    time.sleep(5)
    driver.execute_script("window.scrollTo(0, 800);") 
    time.sleep(10)
     

    save_draft = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Save Draft']")
    save_draft.click()

    time.sleep(5)

    
    

    #social media image upload


def login_bard():
    time.sleep(5)
    driver.get("https://bard.google.com/")
    time.sleep(30)
# Locate and click the Sign In button by its class name
    sign_in_button = driver.find_element(By.XPATH,"//a[@class='gb_va gb_fd gb_Hd gb_ge']")
    sign_in_button.click()
    time.sleep(10)
    # Close the browser when1 done
    # # Input email
    email=wait.until(EC.visibility_of_element_located((By.NAME, 'identifier')))
    email.send_keys("zingtech10@gmail.com")
    time.sleep(2)
    # Click the "Next" button
    next_button_email = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc LQeN7 qIypjc TrZEUc lw1w4b"]')))
    next_button_email.click()
    time.sleep(5)

    # Input password
    password=wait.until(EC.visibility_of_element_located((By.XPATH, '//input[@class="whsOnd zHQkBf"]')))
    password.send_keys("zingtech@123")
    
    time.sleep(10)
    # Click the "Next" button to log in
    next_button_password = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc LQeN7 qIypjc TrZEUc lw1w4b"]')))
    next_button_password.click()
    time.sleep(5)


def login_wordpress1(driver):
    driver.get(WORDPRESS_URL)
    # Implement your login logic for WordPress
    username_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='user_login']")))
    username_field.send_keys(WORD_USERNAME)
    time.sleep(2)
    password_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='user_pass']")))
    password_field.send_keys(WORD_PASSWORD)
    time.sleep(2)

    login_button = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='submit']")))
    login_button.click()
    time.sleep(2)


def fetch_content_from_bard(title):
    driver.get("https://bard.google.com/")
    element = driver.find_element(By.XPATH,"//div[@class='ql-editor ql-blank textarea']")

    # You can interact with the element, for example, send keys or perform other actions
    element.send_keys(f"write a blog with  200 words icluding introduction and conclusion on this topic.-{title}")
    time.sleep(3)
    element.send_keys(Keys.ENTER)
    
    time.sleep(40)
    response_element = WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'response-content')]//div[contains(@class, 'markdown')]")))
    content = response_element.get_attribute("outerHTML")

    # Save the output to a text file
    output_file_path = "C:\\Users\\hp\\Downloads\\blog1.txt"
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        output_file.write(content)

    time.sleep(5) 

    driver.get("https://bard.google.com/")
    time.sleep(5) 
    element2 = driver.find_element(By.XPATH,"//div[@class='ql-editor ql-blank textarea']")

    # You can interact with the element, for example, send keys or perform other actions
    element2.send_keys(f"write a just one word keyphara related to-{title}")
    time.sleep(3)
    element2.send_keys(Keys.ENTER)
    
    time.sleep(40)
    response_element = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'response-content')]//div[contains(@class, 'markdown')]")))
    content = response_element.text

    # Save the output to a text file
    output_file_path = "C:\\Users\\hp\\Downloads\\keys1.txt"
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        output_file.write(content)

   
    time.sleep(5) 

    driver.get("https://bard.google.com/")
    time.sleep(5)
    element3 = driver.find_element(By.XPATH,"//div[@class='ql-editor ql-blank textarea']")

    # You can interact with the element, for example, send keys or perform other actions
    element3.send_keys(f"write top 10 tags headings with single words related to-{title}")
    time.sleep(3)
    element3.send_keys(Keys.ENTER)
    
    time.sleep(40)
    response_element = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'response-content')]//div[contains(@class, 'markdown')]")))
    content = response_element.text

    # Save the output to a text file
    output_file_path = "C:\\Users\\hp\\Downloads\\tags1.txt"
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        output_file.write(content)

    time.sleep(5)
   

   




            



    
  
   

   



if __name__ == "__main__":
    driver.maximize_window()
    
    
    if user_input == '0':
        login_chatgpt()
        login_wordpress(driver)
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):
            title = sheet.cell(row=row, column=2).value
            keyphrase = sheet.cell(row=row, column=3).value
            category = sheet.cell(row=row, column=4).value
            tag_column = sheet.cell(row=row, column=5).value
            imagename = sheet.cell(row=row, column=6).value
            meta_dis=sheet.cell(row=row,column=7).value

            if title:
                print(f"Processing title: {title}")
                try:
                    content = fetch_content_from_chatgpt(title)
                    post_content_to_wordpress(title, content, keyphrase, tag_column, category, imagename,meta_dis)
                except Exception as e:
                    print(f"An error occurred with title {title}: {e}")

    elif user_input == '1':
        driver.maximize_window()
        # Code to execute for another platform
    # Code to execute for another platform
        login_bard()
        login_wordpress1(driver)
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):
            title = sheet.cell(row=row, column=2).value  # Define title here
            keyphrase = sheet.cell(row=row, column=3).value
            category = sheet.cell(row=row, column=4).value
            tag_column = sheet.cell(row=row, column=5).value
            imagename = sheet.cell(row=row, column=6).value
            meta_dis=sheet.cell(row=row,column=7).value

            if title:
                print(f"Processing title: {title}")
                try:
                    # Pass 'title' variable to fetch_content_from_bard function
                    content = fetch_content_from_bard(title)
                    post_content_to_wordpress(title, content, keyphrase, tag_column, category, imagename,meta_dis)
                except Exception as e:
                    print(f"An error occurred with title {title}: {e}")
    else:
        print("Invalid input. Please enter either 0 or 1.")